from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == colName:
            Dict["col"] = i

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict["row"] = i
          
    sheet.cell(row=Dict["row"],column=Dict["col"]).value = new_value
    book.save(file_path)


# Set the path to the chromedriver executable
chromedriver_path = "C:/Users/Usuario/Downloads/chromedriver-win64/chromedriver-win64/chromedriver.exe"

# Set up the service object
service_obj = Service(chromedriver_path)
file_path = "C:/Users/Usuario/Downloads/download.xlsx"
fruit_plate = 'Apple'
new_value = "999"
# Initialize the Chrome driver
driver = webdriver.Chrome(service=service_obj)


driver.implicitly_wait(5)

driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID,"downloadButton").click()

update_excel_data(file_path,"Apple", "price", new_value)

file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text)

price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_plate+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
print(actual_price)
assert actual_price == new_value
